import os

from django.core.wsgi import get_wsgi_application
from openpyxl import load_workbook
from system.utils.configread import config_read
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "admin.settings")
application = get_wsgi_application()
import django
django.setup()

# 数据读取，数据分析处理，并写入DB

# 获取movie目录下的所有.xlsx文件名
from business.models import Category, Industry, Stock

dir_path = 'stock/'
files = [f for f in os.listdir(dir_path) if f.endswith('.xlsx')]

# 存储文件名的变量
category = []

# 遍历文件
for file in files:
    file_path = os.path.join(dir_path, file)

    # 打开Excel文件
    workbook = load_workbook(file_path)
    sheet = workbook.active

    # 获取标题行的内容
    headers = [cell.value for cell in sheet[1]]

    name = file[:-5]
    print(name)

    # 判断标题行的内容是否符合要求
    expected_headers = ['股票名称', '股票代码', '交易所简称', '股票市场类型', '公司名称', '省份', '行业', '公司地址', '公司简介']
    if headers == expected_headers:

        if category is not None:
            # 遍历每一行内容
            for row in sheet.iter_rows(min_row=2, values_only=True):
                # 存储每行内容的变量
                # 分别存储每列的数据
                name, code, address_code, category_name, company, province, industry, company_address, content = row

                # 保存股票市场类型
                stock_category, created = Category.objects.get_or_create(
                    name=category_name
                )
                if created:
                    print("插入股票市场类型成功！")
                else:
                    print('股票市场类型已存在')

                # 保存行业
                last_dash_index = industry.rfind('-')
                if last_dash_index != -1:
                    industry_name = industry[last_dash_index + 1:]

                stock_industry, created = Industry.objects.get_or_create(
                    name=industry_name
                )

                if created:
                    print("插入行业成功！")
                else:
                    print('行业已存在')

                # 保存股票数据
                Stock.objects.create(
                    name=name,
                    content=content,
                    code=code,
                    category_id=stock_category.id,
                    industry_id=stock_industry.id,
                    province=province,
                    address=company_address,
                    company=company,
                    views=1
                )

                print('写入数据成功：'+name)

    else:
        print(f'标题行不符合要求: {file}')

    # 关闭Excel文件
    workbook.close()
