import os

from django.core.wsgi import get_wsgi_application
from openpyxl import load_workbook
from system.utils.configread import config_read

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "admin.settings")
application = get_wsgi_application()
import django

django.setup()

# 数据读取，数据分析处理，并写入DB

# 获取movie目录下的所有.xlsx文件名
from business.models import Stockdata

dir_path = 'stockdata/'
files = [f for f in os.listdir(dir_path) if f.endswith('.xlsx')]

# 存储文件名的变量
category = []

# 遍历文件
for file in files:
    file_path = os.path.join(dir_path, file)

    # 打开Excel文件
    workbook = load_workbook(file_path)
    sheet = workbook.active

    # 获取标题行的内容
    headers = [cell.value for cell in sheet[1]]

    filename = file[:-5]
    print(filename)

    # 判断标题行的内容是否符合要求
    expected_headers = ['股票名称', '股票代码', '最新价', '开盘价', '最低价', '最高价', '成交量', '成交额', '涨跌幅']
    if headers == expected_headers:

        if category is not None:
            # 遍历每一行内容
            for row in sheet.iter_rows(min_row=2, values_only=True):
                # 存储每行内容的变量
                # 分别存储每列的数据
                name, code, newprice, openprice, lowprice, highprice, turnover, amount, chg = row

                if newprice == '-' or openprice == '-' or lowprice == '-' or highprice == '-' or turnover == '-' or amount == '-' or chg == '-':
                    continue

                # 保存股票数据
                Stockdata.objects.create(
                    code=code,
                    newprice=newprice,
                    openprice=openprice,
                    lowprice=lowprice,
                    highprice=highprice,
                    turnover=turnover,
                    amount=amount,
                    time=filename,
                    chg=chg
                )

                print('写入数据成功：' + name)

    else:
        print(f'标题行不符合要求: {file}')

    # 关闭Excel文件
    workbook.close()
